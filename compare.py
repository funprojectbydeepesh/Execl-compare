#compare two excel sheets using python
#too many comments due to this working file being a study by practice 

import openpyxl

#data line paile

data_file1 = openpyxl.load_workbook('tester1.xlsx')
data_file2 = openpyxl.load_workbook('tester2.xlsx')

data_sheet1 = data_file1['Sheet1']
data_sheet2 = data_file2['Sheet1']

#gives title rows
#for this it gives title row [Data, Value]

# for row in data_sheet1.iter_rows():
#     for cell in row:
#         print(cell.value)
#     break

#paile eeuta data herne tarika herum
# print(data_sheet1['A2'].value)

#cell no herne

for row in data_sheet1.iter_rows():
    for cell in row :
        current_cell_value = cell.value
        # print(cell.coordinate)  #gives A1 B1 C1 etc after itering first row.
        cell_locatioin = cell.coordinate
        
        
        # if current_cell_value != data_sheet2[cell_locatioin]:
        #     print(data_sheet2[cell_locatioin])  #this print give A1,B1,A2,B2 ..etc all with data
        if current_cell_value != data_sheet2[cell_locatioin]:
            print(data_sheet2[cell_locatioin])